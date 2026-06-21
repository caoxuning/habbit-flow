"""数据导出路由：打卡记录导出为 Excel"""

import io
from datetime import date
from urllib.parse import quote

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..common import AppException, ok
from ..deps import get_current_user, get_db
from ..models import CheckIn, Goal, User

router = APIRouter(prefix="/api/exports", tags=["导出"])


@router.get("/checkins")
def export_checkins(
    format: str = Query(default="xlsx"),
    start_date: date | None = Query(default=None),
    end_date: date | None = Query(default=None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    导出打卡记录为 Excel 文件
    支持通过 start_date 和 end_date 筛选日期范围
    """
    if format.lower() != "xlsx":
        raise AppException("暂只支持导出 xlsx 格式")

    # 查询数据
    query = (
        db.query(CheckIn, Goal.name, Goal.type)
        .join(Goal, CheckIn.goal_id == Goal.id)
        .filter(CheckIn.user_id == current_user.id)
    )

    if start_date:
        query = query.filter(CheckIn.check_date >= start_date)
    if end_date:
        query = query.filter(CheckIn.check_date <= end_date)

    rows = query.order_by(CheckIn.check_date.desc(), CheckIn.check_time.desc()).all()

    if not rows:
        raise AppException("没有符合条件的打卡记录", 404)

    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "打卡记录"

    # === 样式定义 ===
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F80ED", end_color="2F80ED", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    body_font = Font(name="微软雅黑", size=10)
    body_alignment = Alignment(vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    alt_fill = PatternFill(start_color="F5F8FF", end_color="F5F8FF", fill_type="solid")

    # === 表头 ===
    headers = ["日期", "目标名称", "目标类型", "备注内容", "打卡类型", "打卡时间"]
    col_widths = [14, 24, 14, 45, 12, 20]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # === 数据行 ===
    for row_idx, (checkin, goal_name, goal_type) in enumerate(rows, 2):
        values = [
            checkin.check_date.isoformat(),
            goal_name,
            goal_type,
            checkin.remark or "",
            "补卡" if checkin.makeup else "正常打卡",
            checkin.check_time.strftime("%Y-%m-%d %H:%M"),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            if col_idx in (1, 5, 6):
                cell.alignment = center_alignment
            else:
                cell.alignment = body_alignment

        # 交替行颜色
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill

    # === 设置打印区域和页面 ===
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # === 冻结首行 ===
    ws.freeze_panes = "A2"

    # === 写入内存 ===
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"打卡记录_{date.today().isoformat()}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )
